#!/usr/bin/env python3
"""
Export Excel des données pour PassPrint
Génère des rapports Excel avec openpyxl
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

class ExcelExporter:
    """Exporteur Excel professionnel"""

    def __init__(self):
        self.output_dir = "exports"
        os.makedirs(self.output_dir, exist_ok=True)

    def export_orders(self, orders_data: list) -> str:
        """Exporter les commandes en Excel"""
        filename = f"commandes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Commandes"

        # En-têtes
        headers = ['ID', 'Numéro', 'Client', 'Montant', 'Statut', 'Date', 'Produits']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2D1B69", end_color="2D1B69", fill_type="solid")

        # Données
        for row, order in enumerate(orders_data, 2):
            ws.cell(row=row, column=1, value=order.get('id', ''))
            ws.cell(row=row, column=2, value=order.get('order_number', ''))
            ws.cell(row=row, column=3, value=order.get('customer_name', ''))
            ws.cell(row=row, column=4, value=order.get('total_amount', 0))
            ws.cell(row=row, column=5, value=order.get('status', ''))
            ws.cell(row=row, column=6, value=order.get('created_at', ''))
            ws.cell(row=row, column=7, value=str(order.get('items', [])))

        # Ajuster la largeur des colonnes
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

        wb.save(filepath)
        return filepath

    def export_products(self, products_data: list) -> str:
        """Exporter les produits en Excel"""
        filename = f"produits_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produits"

        # En-têtes
        headers = ['ID', 'Nom', 'Prix', 'Catégorie', 'Stock', 'Statut']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")

        # Données
        for row, product in enumerate(products_data, 2):
            ws.cell(row=row, column=1, value=product.get('id', ''))
            ws.cell(row=row, column=2, value=product.get('name', ''))
            ws.cell(row=row, column=3, value=product.get('price', 0))
            ws.cell(row=row, column=4, value=product.get('category', ''))
            ws.cell(row=row, column=5, value=product.get('stock_quantity', 0))
            ws.cell(row=row, column=6, value='Actif' if product.get('is_active') else 'Inactif')

        # Ajuster la largeur des colonnes
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15

        wb.save(filepath)
        return filepath

# Instance globale de l'exporteur Excel
excel_exporter = ExcelExporter()

def export_orders_to_excel(orders_data: list) -> str:
    """Exporter les commandes en Excel"""
    return excel_exporter.export_orders(orders_data)

def export_products_to_excel(products_data: list) -> str:
    """Exporter les produits en Excel"""
    return excel_exporter.export_products(products_data)

if __name__ == "__main__":
    print("Système d'export Excel opérationnel!")